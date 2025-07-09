"""
Excel file processor for reading and validating data
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelProcessor:
    """Process Excel files for conversion"""
    
    REQUIRED_COLUMNS = ['matricula', 'rubrica', 'valor', 'tipo', 'trigrama']
    VALID_TYPES = ['NO', 'DE']
    
    def __init__(self):
        pass
        
    def process_file(self, file_path):
        """Process Excel file and return validated data"""
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Validate columns
            self._validate_columns(df)
            
            # Clean and validate data
            processed_data = []
            for index, row in df.iterrows():
                record = self._process_record(row, index + 1)
                if record:
                    processed_data.append(record)
                    
            return processed_data
            
        except Exception as e:
            raise Exception(f"Erro ao processar arquivo Excel: {str(e)}")
            
    def _validate_columns(self, df):
        """Validate required columns exist"""
        missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
        if missing_columns:
            raise Exception(f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}")
            
    def _process_record(self, row, line_number):
        """Process individual record"""
        try:
            # Extract and clean data
            matricula = str(row['matricula']).strip()
            rubrica = str(row['rubrica']).strip()
            valor = str(row['valor']).strip()
            tipo = str(row['tipo']).strip().upper()
            trigrama = str(row['trigrama']).strip().upper()
            
            # Validate required fields
            if not matricula or matricula == 'nan':
                raise ValueError("Matrícula é obrigatória")
                
            if not rubrica or rubrica == 'nan':
                raise ValueError("Rubrica é obrigatória")
                
            if not valor or valor == 'nan':
                raise ValueError("Valor é obrigatório")
                
            # Validate and format matricula (should be numeric)
            if not matricula.isdigit():
                raise ValueError("Matrícula deve conter apenas números")
                
            # Validate and format rubrica (should be 7 digits)
            if not rubrica.isdigit() or len(rubrica) != 7:
                raise ValueError("Rubrica deve conter 7 dígitos")
                
            # Validate and format valor (convert to float)
            valor = valor.replace(',', '.')  # Handle comma as decimal separator
            try:
                valor_float = float(valor)
                valor_formatted = f"{valor_float:.2f}"
            except ValueError:
                raise ValueError("Valor deve ser um número válido")
                
            # Validate tipo
            if tipo not in self.VALID_TYPES:
                raise ValueError(f"Tipo deve ser NO ou DE, encontrado: {tipo}")
                
            # Validate trigrama (should be 3 characters)
            if len(trigrama) != 3:
                raise ValueError("Trigrama deve conter 3 caracteres")
                
            return {
                'matricula': matricula,
                'rubrica': rubrica,
                'valor': valor_formatted,
                'tipo': tipo,
                'trigrama': trigrama,
                'valid': True,
                'line_number': line_number
            }
            
        except Exception as e:
            # Return invalid record with error info
            return {
                'matricula': str(row.get('matricula', '')),
                'rubrica': str(row.get('rubrica', '')),
                'valor': str(row.get('valor', '')),
                'tipo': str(row.get('tipo', '')),
                'trigrama': str(row.get('trigrama', '')),
                'valid': False,
                'error': str(e),
                'line_number': line_number
            }
            
    def create_template(self, file_path):
        """Create Excel template with example data and instructions"""
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create Comandos sheet
        ws_comandos = wb.active
        ws_comandos.title = "Comandos"
        
        # Headers
        headers = ['matricula', 'rubrica', 'valor', 'tipo', 'trigrama']
        for col, header in enumerate(headers, 1):
            cell = ws_comandos.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        # Example data
        example_data = [
            ['10024450', '1208000', '12000.00', 'NO', 'BAA'],
            ['10024450', '1210005', '12000.00', 'DE', 'BAA'],
            ['97115215', '1208000', '3066.09', 'NO', 'BAA'],
            ['97115215', '1213101', '3066.09', 'DE', 'BAA'],
            ['98004450', '1208000', '21944.53', 'NO', 'BAA']
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws_comandos.cell(row=row, column=col, value=value)
                
        # Format columns
        ws_comandos.column_dimensions['A'].width = 12  # matricula
        ws_comandos.column_dimensions['B'].width = 10  # rubrica
        ws_comandos.column_dimensions['C'].width = 12  # valor
        ws_comandos.column_dimensions['D'].width = 6   # tipo
        ws_comandos.column_dimensions['E'].width = 10  # trigrama
        
        # Add data validation for tipo column
        dv = DataValidation(type="list", formula1='"NO,DE"', allow_blank=False)
        dv.error = "Valor deve ser NO ou DE"
        dv.errorTitle = "Valor Inválido"
        ws_comandos.add_data_validation(dv)
        dv.add(f"D2:D1000")
        
        # Create Instructions sheet
        ws_instructions = wb.create_sheet("Instruções")
        
        instructions = [
            ["INSTRUÇÕES DE USO", ""],
            ["", ""],
            ["1. COLUNAS OBRIGATÓRIAS:", ""],
            ["   • matricula", "Número da matrícula do funcionário (apenas números)"],
            ["   • rubrica", "Código da rubrica (7 dígitos)"],
            ["   • valor", "Valor monetário (usar ponto ou vírgula como separador decimal)"],
            ["   • tipo", "Tipo do lançamento: NO (normal) ou DE (desconto)"],
            ["   • trigrama", "Código do trigrama (3 caracteres)"],
            ["", ""],
            ["2. EXEMPLOS DE DADOS VÁLIDOS:", ""],
            ["   • Matrícula: 10024450, 97115215, 98004450"],
            ["   • Rubrica: 1208000, 1210005, 1213101"],
            ["   • Valor: 12000.00, 3066.09, 21944.53"],
            ["   • Tipo: NO, DE"],
            ["   • Trigrama: BAA, XYZ, ABC"],
            ["", ""],
            ["3. REGRAS IMPORTANTES:", ""],
            ["   • Não deixe células vazias nas colunas obrigatórias"],
            ["   • Matrícula deve conter apenas números"],
            ["   • Rubrica deve ter exatamente 7 dígitos"],
            ["   • Valores devem ser números válidos"],
            ["   • Tipo deve ser exatamente NO ou DE"],
            ["   • Trigrama deve ter exatamente 3 caracteres"],
            ["", ""],
            ["4. DICAS:", ""],
            ["   • Use os exemplos da aba 'Comandos' como referência"],
            ["   • Teste com poucos registros antes de processar arquivo completo"],
            ["   • Verifique se todas as colunas estão preenchidas"],
            ["   • Mantenha o formato original das colunas"]
        ]
        
        for row, (title, desc) in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=title)
            ws_instructions.cell(row=row, column=2, value=desc)
            
            # Format titles
            if title and not title.startswith("   "):
                cell = ws_instructions.cell(row=row, column=1)
                cell.font = Font(bold=True)
                if title.startswith("INSTRUÇÕES"):
                    cell.font = Font(bold=True, size=14)
                    
        # Format instructions sheet
        ws_instructions.column_dimensions['A'].width = 30
        ws_instructions.column_dimensions['B'].width = 50
        
        # Save workbook
        wb.save(file_path)
